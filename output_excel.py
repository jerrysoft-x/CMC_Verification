from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors, Alignment
from openpyxl.comments import Comment
from certificate_element import CertificateElementToVerify, CertificateElementInPlate, ChemicalElementValue
from common_utils import Certificate


def output_title(sheet, row_cursor: int, element: CertificateElementToVerify):
    bold_font = Font(bold=True)
    redFill = PatternFill(
        start_color='FFFF0000',
        end_color='FFFF0000',
        fill_type='solid'
    )
    sheet.cell(row=row_cursor, column=1).value = element.__class__.__name__.upper()
    sheet.cell(row=row_cursor, column=1).font = bold_font
    sheet.cell(row=row_cursor, column=2).value = element.value
    if not element.valid_flag:
        sheet.cell(row=row_cursor, column=2).fill = redFill
        sheet.cell(row=row_cursor, column=2).comment = Comment(element.message, 'CMC_Verification')


def output_value(sheet, row_cursor: int, column_cursor: int, element: CertificateElementInPlate):
    normal_font = Font(bold=False)
    redFill = PatternFill(
        start_color='FFFF0000',
        end_color='FFFF0000',
        fill_type='solid'
    )
    center_aligned_text = Alignment(horizontal="center")
    sheet.cell(row=row_cursor, column=column_cursor).font = normal_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    if isinstance(element, ChemicalElementValue):
        sheet.cell(row=row_cursor, column=column_cursor).value = element.calculated_value
    else:
        sheet.cell(row=row_cursor, column=column_cursor).value = element.value
    if isinstance(element, CertificateElementToVerify) and not element.valid_flag:
        sheet.cell(row=row_cursor, column=column_cursor).fill = redFill
        sheet.cell(row=row_cursor, column=column_cursor).comment = Comment(element.message, 'CMC_Verification')


def output_excel(certificate: Certificate, input_file: str, output_file: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = input_file.replace('.pdf', '')
    row_cursor = 1
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    center_aligned_text = Alignment(horizontal="center")
    # Steel Plant
    sheet.cell(row=row_cursor, column=1).value = 'STEEL PLANT'
    sheet.cell(row=row_cursor, column=1).font = bold_font
    sheet.cell(row=row_cursor, column=2).value = certificate.steel_plant
    row_cursor += 1
    # Specification
    output_title(sheet, row_cursor, certificate.specification)
    row_cursor += 1
    # Thickness
    output_title(sheet, row_cursor, certificate.thickness)
    row_cursor += 1

    # Steel Plates
    row_cursor += 1
    column_cursor = 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'STEEL PLATES'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'MATERIAL DESCRIPTION'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'CHEMICAL COMPOSITION %'
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + len(certificate.chemical_elements) - 1)
    column_cursor += len(certificate.chemical_elements)
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'TENSILE TEST'
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + 3 - 1)
    column_cursor += 3
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'IMPACT TEST'
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + 6 - 1)
    column_cursor += 6
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'DELIVERY CONDITION'
    column_cursor += 1

    row_cursor += 1
    column_cursor = 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'NO.'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'MASS(kg)'
    column_cursor += 1
    for index, element in enumerate(certificate.chemical_elements):
        sheet.cell(row=row_cursor, column=column_cursor + index).font = bold_font
        sheet.cell(row=row_cursor, column=column_cursor + index).alignment = center_aligned_text
        sheet.cell(row=row_cursor, column=column_cursor + index).value = element
    column_cursor += len(certificate.chemical_elements)
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'Y.S.'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'T.S.'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'EL'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'POSITION DIRECTION'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'TEMPERATURE'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = '1'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = '2'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = '3'
    column_cursor += 1
    sheet.cell(row=row_cursor, column=column_cursor).font = bold_font
    sheet.cell(row=row_cursor, column=column_cursor).alignment = center_aligned_text
    sheet.cell(row=row_cursor, column=column_cursor).value = 'AVE.'
    column_cursor += 1

    row_cursor += 1
    for plate_index, steel_plate in enumerate(certificate.steel_plates):
        column_cursor = 1
        # No.
        sheet.cell(row=row_cursor + plate_index, column=column_cursor).font = normal_font
        sheet.cell(row=row_cursor + plate_index, column=column_cursor).alignment = center_aligned_text
        sheet.cell(row=row_cursor + plate_index, column=column_cursor).value = steel_plate.serial_number
        column_cursor += 1
        # MASS
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.mass)
        column_cursor += 1
        # CHEMICAL COMPOSITION
        for elem_index, elem in enumerate(steel_plate.chemical_compositions):
            output_value(sheet, row_cursor + plate_index, column_cursor + elem_index,
                         steel_plate.chemical_compositions[elem])
        column_cursor += len(steel_plate.chemical_compositions)
        # TENSILE TEST
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.yield_strength)
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.tensile_strength)
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.elongation)
        column_cursor += 1
        # IMPACT TEST
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.position_direction_impact)
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.temperature)
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.impact_energy_list[0])
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.impact_energy_list[1])
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.impact_energy_list[2])
        column_cursor += 1
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.impact_energy_list[3])
        column_cursor += 1
        # DELIVERY CONDITION
        output_value(sheet, row_cursor + plate_index, column_cursor, steel_plate.delivery_condition)
        column_cursor += 1

    workbook.save(filename=output_file)
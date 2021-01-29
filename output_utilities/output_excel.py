import os
from typing import Any, Tuple, Union, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from common import Certificate, CommonUtils, CertificateElementToVerify, CertificateElementInPlate, \
    ChemicalElementValue, PositionDirectionImpact

normal_font = Font(bold=False)
bold_font = Font(bold=True)
left_aligned_text = Alignment(horizontal="left")
center_aligned_text = Alignment(horizontal="center")
redFill = PatternFill(
    start_color='FFFF0000',
    end_color='FFFF0000',
    fill_type='solid'
)
thick = Side(border_style="thick", color="000000")
thick_outsides_border = Border(top=thick, bottom=thick, left=thick, right=thick)

# def output_title(sheet, row_cursor: int, element: CertificateElementToVerify):
#     sheet.cell(row=row_cursor, column=1).value = element.__class__.__name__.upper()
#     sheet.cell(row=row_cursor, column=1).font = bold_font
#     sheet.cell(row=row_cursor, column=2).value = element.value
#     if not element.valid_flag:
#         sheet.cell(row=row_cursor, column=2).fill = redFill
#         sheet.cell(row=row_cursor, column=2).comment = Comment(element.message, 'CMC_Verification')


def write_title(sheet, row_cursor: int, column_cursor: int, value: Any, font: Font = bold_font,
                alignment: Alignment = center_aligned_text, border: Border = thick_outsides_border):
    cell = sheet.cell(row_cursor, column_cursor)
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    cell.border = border


def write_value(sheet, row_cursor: int, column_cursor: int, element: Union[CertificateElementInPlate, str],
                font: Font = normal_font, alignment: Alignment = left_aligned_text):
    cell = sheet.cell(row=row_cursor, column=column_cursor)
    cell.font = font
    cell.alignment = alignment
    if isinstance(element, str):
        cell.value = element
    else:
        if isinstance(element, ChemicalElementValue):
            cell.value = element.calculated_value
        elif isinstance(element, PositionDirectionImpact):
            cell.value = str(element.value)
        else:
            cell.value = element.value
        if isinstance(element, CertificateElementToVerify) and not element.valid_flag:
            cell.fill = redFill
            cell.comment = Comment(element.message, 'CMC_Verification')


def initialize_workbook(sheet_name: str) -> Tuple[Workbook, Worksheet, int, int]:
    # Initialize excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    row_cursor, column_cursor = 1, 1

    # Prepare the title area
    # [SECTION] HEAD
    write_title(sheet, row_cursor, column_cursor, 'HEAD')
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + 2)
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, 'FILE NAME')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'STEEL PLANT')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'CERTIFICATE NO.')
    # [SECTION] STEEL PLATES
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'STEEL PLATES')
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + 1)
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, 'BATCH NO.')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'PLATE NO.')
    # [SECTION] MATERIAL DESCRIPTION
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'MATERIAL DESCRIPTION')
    sheet.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=column_cursor,
                      end_column=column_cursor + 3)
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, 'SPECIFICATION')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'THICKNESS')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'QUANTITY')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'MASS(ton)')
    # [SECTION] CHEMICAL COMPOSITION
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'CHEMICAL COMPOSITION')
    sheet.merge_cells(
        start_row=row_cursor,
        end_row=row_cursor,
        start_column=column_cursor,
        end_column=column_cursor + len(CommonUtils.chemical_elements_table) - 1
    )
    row_cursor += 1
    column_cursor -= 1
    for element in CommonUtils.chemical_elements_table:
        write_title(sheet, row_cursor, column_cursor := column_cursor + 1, element)
    # [SECTION] TENSILE TEST
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'TENSILE TEST')
    sheet.merge_cells(
        start_row=row_cursor,
        end_row=row_cursor,
        start_column=column_cursor,
        end_column=column_cursor + 2
    )
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, 'Y.S.')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'T.S.')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'EL')
    # [SECTION] IMPACT TEST
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'IMPACT TEST')
    sheet.merge_cells(
        start_row=row_cursor,
        end_row=row_cursor,
        start_column=column_cursor,
        end_column=column_cursor + 5
    )
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, 'POSITION DIRECTION')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'TEMPERATURE')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, '1')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, '2')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, '3')
    write_title(sheet, row_cursor, column_cursor := column_cursor + 1, 'AVE.')
    # [SECTION] DELIVERY CONDITION
    write_title(sheet, row_cursor := row_cursor - 1, column_cursor := column_cursor + 1, 'DELIVERY CONDITION')
    write_title(sheet, row_cursor := row_cursor + 1, column_cursor, '')

    row_cursor += 1
    column_cursor = 1

    return workbook, sheet, row_cursor, column_cursor

    # print(row_cursor, column_cursor)

    # workbook.save(filename=output_file)


def write_single_certificate(certificate: Certificate, sheet, row_cursor: int) -> int:
    for plate_index, steel_plate in enumerate(certificate.steel_plates):
        # Reset column cursor
        column_cursor = 1
        # [SECTION] HEAD
        #   [CELL] FILE NAME
        write_value(sheet, row_cursor + plate_index, column_cursor, certificate.file_path)
        #   [CELL] STEEL PLANT
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, certificate.steel_plant)
        #   [CELL] CERTIFICATE NO
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, certificate.certificate_no)
        # [SECTION] STEEL PLATES
        #   [CELL] BATCH NO.
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.batch_no)
        #   [CELL] PLATE NO.
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.plate_no)
        # [SECTION] MATERIAL DESCRIPTION
        #   [CELL] SPECIFICATION
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.specification)
        #   [CELL] THICKNESS
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.thickness)
        #   [CELL] QUANTITY
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.quantity)
        #   [CELL] MASS(kg)
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.mass)
        # [SECTION] CHEMICAL COMPOSITION
        #   [CELL] EACH ELEMENT
        for elem in CommonUtils.chemical_elements_table:
            if elem in steel_plate.chemical_compositions:
                write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                            steel_plate.chemical_compositions[elem])
            else:
                write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, 'N/A')
        # [SECTION] TENSILE TEST
        #   [CELL] Y.S.
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.yield_strength)
        #   [CELL] T.S.
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.tensile_strength)
        #   [CELL] EL
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.elongation)
        # [SECTION] IMPACT TEST
        #   [CELL] POSITION DIRECTION
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                    steel_plate.position_direction_impact)
        #   [CELL] TEMPERATURE
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1, steel_plate.temperature)
        #   [CELL] IMPACT ENERGY 1
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                    steel_plate.impact_energy_list[0] if len(steel_plate.impact_energy_list) == 4 else '')
        #   [CELL] IMPACT ENERGY 2
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                    steel_plate.impact_energy_list[1] if len(steel_plate.impact_energy_list) == 4 else '')
        #   [CELL] IMPACT ENERGY 3
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                    steel_plate.impact_energy_list[2] if len(steel_plate.impact_energy_list) == 4 else '')
        #   [CELL] IMPACT ENERGY AVE.
        write_value(sheet, row_cursor + plate_index, column_cursor := column_cursor + 1,
                    steel_plate.impact_energy_list[3] if len(steel_plate.impact_energy_list) == 4 else '')
        # [SECTION] DELIVERY CONDITION
        write_value(sheet, row_cursor + plate_index, column_cursor + 1, steel_plate.delivery_condition)

    return row_cursor + len(certificate.steel_plates)


def write_single_certificate_to_excel(certificate: Certificate, sheet_name: str, output_file: str):
    workbook, sheet, row_cursor, column_cursor = initialize_workbook(sheet_name)
    write_single_certificate(certificate, sheet, row_cursor)
    workbook.save(filename=output_file)


def write_multiple_certificates_to_excel(certificates: List[Certificate], sheet_name: str = 'PASS',
                                         output_file: str = os.path.join('PASS', 'PASS.xlsx')):
    workbook, sheet, row_cursor, column_cursor = initialize_workbook(sheet_name)
    for certificate in certificates:
        row_cursor = write_single_certificate(certificate, sheet, row_cursor)
    workbook.save(filename=output_file)


def write_certificates_with_exception(certificates_with_exception: List[Tuple[str, str]], sheet_name: str = 'EXCEPTION',
                                      output_file: str = os.path.join('EXCEPTION', 'EXCEPTION.xlsx')):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    write_title(sheet, row_cursor := 1, column_cursor := 1, 'FILE NAME')
    write_title(sheet, row_cursor, column_cursor + 1, 'EXCEPTION MESSAGE')
    for file_name, exception_message in certificates_with_exception:
        write_value(sheet, row_cursor := row_cursor + 1, column_cursor := 1, file_name)
        write_value(sheet, row_cursor, column_cursor + 1, exception_message)
    workbook.save(filename=output_file)


if __name__ == '__main__':
    # output_single_certificate_to_excel(None, 'Test', 'test.xlsx')
    pass
